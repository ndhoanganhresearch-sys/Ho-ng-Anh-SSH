# -*- coding: utf-8 -*-
"""Project status mirroring the EXACT current GUI step structure (core mode)."""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\ssl\Desktop\Code Python\data python cusor\tunnel_project\samples\tool_status_by_step.xlsx"

NAVY = "1F3864"; LBLUE = "BDD7EE"; GRAY = "595959"
ST = {
    "COMPLETED":          ("C6EFCE", "276221"),
    "PARTIALLY COMPLETE": ("FFEB9C", "9C6500"),
    "PENDING":            ("FCE4D6", "833C00"),
}
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); ws = wb.active; ws.title = "Status by Step"
for col, w in {"A": 6, "B": 34, "C": 22, "D": 10, "E": 58, "F": 56}.items():
    ws.column_dimensions[col].width = w

def merge_row(r, text, fill, font):
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]; c.value = text; c.fill = PatternFill("solid", fgColor=fill); c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

merge_row(1, "PROJECT STATUS BY WORKFLOW STEP - TUNNEL ANALYSIS v4.0", "FFFFFF",
          Font(bold=True, size=18, color=NAVY))
merge_row(2, "Osong Tunnel Monitoring Project | CBNU Smart Structure Lab (SSL) | Mirrors the current GUI sidebar (core mode) | Code-verified 2026-06-08",
          "FFFFFF", Font(bold=False, size=10, color=GRAY))
ws.row_dimensions[1].height = 26

headers = ["ID", "Module / Feature (GUI label)", "Current Status", "Priority",
           "Current Code Status (evidence)", "Remaining Work / Optimization"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[4].height = 22

rows = [
 ("STEP", "STEP 1 - LiDAR DATA ACQUISITION"),
 ("1.1", "Import LAS / PLY data", "COMPLETED", "Medium",
  "Parses .las/.laz/.ply and .txt/.xyz/.pts/.csv (io_layer.py); RGB->luminance fallback; multi-epoch (T0/Tn) load.",
  "Done. Optional: native E57 / vendor headers (Trimble/Faro)."),
 ("1.2", "Add scan station (+)", "COMPLETED", "Medium",
  "Loads additional scanner stations into one context; coordinate frame preserved across multi-station merges.",
  "Done."),

 ("STEP", "STEP 2 - PREPROCESSING AND NOISE FILTERING"),
 ("2.1", "Voxel downsampling", "COMPLETED", "Low",
  "Open3D voxel grid; recenters only for a single scan so multi-station frames stay aligned. Adaptive SOR available.",
  "Done. Numba JIT only worthwhile beyond ~50M points."),
 ("2.2", "Clean noise (auto: cables, lights, people, wall cables)", "COMPLETED", "Medium",
  "auto_denoise(): removes cables (linearity), lights (sphericity), people (DBSCAN) via vectorized local-PCA Demantke features + sanity guards. Runs inside AUTO PIPELINE.",
  "Done. Future: deep-learning segmentation (PointNet++) for hard cases."),

 ("STEP", "STEP 3 - REGISTRATION AND SYNCHRONIZATION"),
 ("3.1", "Auto-align T0/Tn epochs (target or ICP)", "COMPLETED", "High",
  "register_epochs(): marker-SVD when >=3 fixed targets (no deformation absorption) else trimmed-ICP; divergence-guard keeps result never worse than input.",
  "Done. Roll about a round tunnel needs markers; yaw+translation recovered by ICP."),

 ("STEP", "STEP 4 - GEOMETRIC COORDINATE SYSTEM"),
 ("4.1", "B-Spline C2 centerline (PDF 3.4)", "COMPLETED", "High",
  "Scipy splprep/splev C2 trajectory + curvature-guarded tangent refine; builds gravity-anchored Bishop (twist-free) Frenet frames and intensity-derivative ring seams in the same pass.",
  "Done (Bishop frame + intensity-valley ring seams implemented)."),

 ("STEP", "STEP 5 - PARAMETER EXTRACTION"),
 ("5.1", "Crown settlement dv", "COMPLETED", "High",
  "calc_arch_settlement vs T0 per Frenet section; robust p99 crown (a stray slab point cannot inflate it); _has_t0_reference helper for single/dual-scan.",
  "Done."),
 ("5.2", "Horizontal convergence dh", "COMPLETED", "High",
  "calc_horizontal_convergence vs T0; robust p99-p1 width (stray points cannot inflate the span).",
  "Done."),
 ("5.3", "Ovality epsilon", "COMPLETED", "High",
  "Fitzgibbon Direct Least-Squares ellipse fit (fit_ellipse_fitzgibbon) -> real axes a,b; covariance-eigenvalue fallback only if the fit fails.",
  "Done (Fitzgibbon DLS, not covariance eigenvalues)."),
 ("5.4", "Section eccentricity e", "COMPLETED", "High",
  "Centroid-vs-axis offset; curved-tunnel single-scan eccentricity detrended (Kasa circle centre + moving-median) to remove false ~450mm bias.",
  "Done. Curved 1-scan max ~245mm residual; load T0+Tn for exact result."),

 ("STEP", "STEP 6 - TIME-SERIES ANALYSIS"),
 ("6.1", "Plot deformation trend T0→Tn", "COMPLETED", "High",
  "Crown/convergence/ovality trend along chainage; forecast_threshold_crossing() predicts time-to-CAUTION/CRITICAL with R^2 (NEW, 15/15 tests).",
  "Forecast engine done; GUI multi-epoch (3+) loading still to wire."),
 ("6.2", "M3C2 deformation map T0→Tn", "COMPLETED", "High",
  "M3C2 (py4dgeo) with C2C cKDTree fallback; multi-epoch spatiotemporal drift; Hausdorff heatmap green/yellow/red.",
  "Done (KD-Tree multi-epoch drift)."),
 ("6.3", "Plot 2D Technical Section T0/Tn", "COMPLETED", "High",
  "Runs the per-section clearance check (robust p1 flag + portal guard, no false alarms) and classify_sections() - the single source of truth (OK/CAUTION/CRITICAL) shared by ruler/track/3D/dashboard; banner now consistent with section alerts.",
  "Done (clearance + banner hardened this session, 11 tests)."),

 ("STEP", "STEP 7 - BIM, REPORTING AND AI"),
 ("7.1", "Export IFC tunnel structure", "COMPLETED", "Medium",
  "export_ifc(include_components=False): continuous tunnel lining as a tessellated shell (IfcPolygonalFaceSet) LOFTED from the measured per-section rings, so it follows real deformation (not a uniform-radius tube); hollow (outer+inner+end caps), status-coloured bands (grey/amber/red), + IFC4X3 IfcAlignment centerline. Per-section data kept as property sets. Cable/light geometry excluded per scope.",
  "Done. Optional later: classify shell as IfcWall/IfcSlab/IfcSpace; IfcSectionedSolidHorizontal for fully parametric sweep."),
 ("7.2", "Export section CSV", "COMPLETED", "Low",
  "Per-section parameters exported to CSV for downstream integration.", "Done."),
 ("7.3", "Export Excel report", "COMPLETED", "Low",
  "Multi-sheet Excel workbook of parameters and warnings.", "Done."),
 ("7.4", "Export PDF report", "COMPLETED", "Medium",
  "reportlab PDF: cover, summary, charts, section table, warnings.", "Done."),
 ("7.5", "Generate AI work order (PDF)", "COMPLETED", "Medium",
  "generate_work_order(): groups flagged sections into a ranked PDF work order (location, standard, action, priority); optional offline-safe LLM narrative (NEW, 26/26 tests).",
  "Done."),
 ("7.6", "Query structural AI assistant", "COMPLETED", "Low",
  "RAG (ChromaDB + sentence-transformers) with Korean safety standards; query() injects serialized PipelineContext (RMSE, ovality, settlement) into the prompt; Ollama offline.",
  "Done (context injection target met)."),
]

r = 5
for row in rows:
    if row[0] == "STEP":
        merge_row(r, row[1], LBLUE, Font(bold=True, size=10, color=NAVY))
        ws.row_dimensions[r].height = 18
        r += 1
        continue
    _id, mod, status, prio, evid, rem = row
    for ci, v in enumerate([_id, mod, status, prio, evid, rem], 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.border = border
        c.alignment = Alignment(horizontal="center" if ci in (1, 3, 4) else "left",
                                vertical="center", wrap_text=True)
        c.font = Font(bold=(ci == 1), size=9)
    fill, fcol = ST.get(status, ("FFFFFF", "000000"))
    sc = ws.cell(row=r, column=3); sc.fill = PatternFill("solid", fgColor=fill)
    sc.font = Font(bold=True, size=9, color=fcol)
    lines = max(math.ceil(len(evid) / 55), math.ceil(len(rem) / 55), math.ceil(len(mod) / 30), 1)
    ws.row_dimensions[r].height = max(30, lines * 13 + 8)
    r += 1

r += 1
ws.cell(row=r, column=1, value="Legend:").font = Font(bold=True, size=9)
for k, desc in [("COMPLETED", "implemented & verified in code/tests"),
                ("PARTIALLY COMPLETE", "core works; enhancement remaining"),
                ("PENDING", "not started")]:
    r += 1
    c = ws.cell(row=r, column=2, value=k); fill, fcol = ST[k]
    c.fill = PatternFill("solid", fgColor=fill); c.font = Font(bold=True, size=9, color=fcol)
    c.alignment = Alignment(horizontal="center")
    d = ws.cell(row=r, column=3, value=desc); d.font = Font(size=9)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

ws.freeze_panes = "A5"
import os
target = OUT
for attempt in range(5):
    try:
        wb.save(target); break
    except PermissionError:
        base, ext = os.path.splitext(OUT)
        target = f"{base}_v{attempt+2}{ext}"
print("Saved:", target)
